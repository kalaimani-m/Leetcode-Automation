"""
Daily LeetCode student progress report.

Usage:
    python main.py            # normal daily run -> LeetCode_Daily_Report_YYYY-MM-DD.xlsx
    python main.py --test     # test run -> LeetCode_Daily_Report_TEST.xlsx, no email sent

Reads students/leetcode_Links.xlsx (the master list), queries each student's
public LeetCode profile, writes the report, appends to data/history.csv,
and (unless --test) emails the report.
"""

import os
import sys
import csv
import json
import time
import smtplib
import argparse
from email.message import EmailMessage
from datetime import datetime
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from leetcode_client import extract_username, fetch_profile, count_unique_solved_today

IST = ZoneInfo("Asia/Kolkata")
STUDENTS_FILE = os.path.join(os.path.dirname(__file__), "students", "leetcode_Links.xlsx")
REPORTS_DIR = os.path.join(os.path.dirname(__file__), "reports")
HISTORY_FILE = os.path.join(os.path.dirname(__file__), "data", "history.csv")
DELAY_BETWEEN_STUDENTS = 2  # seconds -- be polite to LeetCode, avoid rate-limit bans

RECEIVER_EMAIL = "kalaimani.cybersec@prathyusha.edu.in"


def load_students(path):
    """Find the real header row (S.No / Reg.no / ...) and return a clean DataFrame."""
    raw = pd.read_excel(path, header=None)
    header_row_idx = None
    for i in range(min(5, len(raw))):
        row_values = [str(v).strip().lower() for v in raw.iloc[i].tolist()]
        if any("s.no" in v for v in row_values):
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError("Could not find the header row (looking for 'S.No') in the master Excel file.")

    df = pd.read_excel(path, header=header_row_idx)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(subset=[df.columns[0]])  # drop fully blank trailing rows
    return df


def process_students(df, today_ist_date):
    results = []
    total = len(df)

    for i, row in df.reset_index(drop=True).iterrows():
        sno = row.get("S.No")
        regno = row.get("Reg.no")
        name = row.get("Name of the Student")
        url = row.get("link")

        username = extract_username(url) if isinstance(url, str) else None

        if not username:
            results.append({
                "sno": sno, "regno": regno, "name": name,
                "today": "N/A", "overall": "N/A",
                "note": "No usable profile URL in master file",
            })
            print(f"[{i+1}/{total}] {name}: SKIPPED (no profile URL)")
            continue

        profile = fetch_profile(username)

        if profile["status"] != "ok":
            results.append({
                "sno": sno, "regno": regno, "name": name,
                "today": "N/A", "overall": "N/A",
                "note": f"Profile unavailable ({profile['status']}: {profile['error']})",
            })
            print(f"[{i+1}/{total}] {name}: N/A ({profile['status']})")
        else:
            today_count, capped = count_unique_solved_today(profile["recent_ac"], today_ist_date)
            overall = profile["overall_solved"] if profile["overall_solved"] is not None else "N/A"
            note = ""
            if capped:
                note = "Recent-activity window is full (20 entries) -- today's count may be a lower bound"
            results.append({
                "sno": sno, "regno": regno, "name": name,
                "today": today_count if today_count is not None else "N/A",
                "overall": overall,
                "note": note,
            })
            print(f"[{i+1}/{total}] {name}: today={today_count} overall={overall}")

        time.sleep(DELAY_BETWEEN_STUDENTS)

    return results


def build_report(results, today_ist_date, out_path):
    wb = Workbook()

    ws = wb.active
    ws.title = "Daily Report"
    headers = ["S.No.", "Registration Number", "Student Name", "Current Date",
               "Problems Solved on Current Date", "Overall Problems Solved"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")

    date_str = today_ist_date.strftime("%d-%b-%Y")
    for r in results:
        ws.append([r["sno"], r["regno"], r["name"], date_str, r["today"], r["overall"]])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    widths = [6, 20, 28, 14, 28, 22]
    for col, w in zip("ABCDEF", widths):
        ws.column_dimensions[col].width = w

    # Second sheet: anything worth the user's attention -- doesn't touch the
    # required 6-column layout above, it's supplementary.
    notes = [r for r in results if r["note"]]
    if notes:
        ws2 = wb.create_sheet("Notes")
        ws2.append(["Student Name", "Note"])
        for cell in ws2[1]:
            cell.font = Font(name="Arial", bold=True)
        for r in notes:
            ws2.append([r["name"], r["note"]])
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 70

    wb.save(out_path)


def append_history(results, today_ist_date):
    os.makedirs(os.path.dirname(HISTORY_FILE), exist_ok=True)
    file_exists = os.path.isfile(HISTORY_FILE)
    with open(HISTORY_FILE, "a", newline="") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["Date", "Reg.no", "Student Name", "Problems Solved Today", "Overall Solved"])
        date_str = today_ist_date.strftime("%d-%b-%Y")
        for r in results:
            writer.writerow([date_str, r["regno"], r["name"], r["today"], r["overall"]])


def send_email(report_path, today_ist_date):
    sender = os.environ.get("SENDER_EMAIL")
    app_password = os.environ.get("SENDER_APP_PASSWORD")
    if not sender or not app_password:
        print("SENDER_EMAIL / SENDER_APP_PASSWORD not set -- skipping email send.")
        return

    date_str = today_ist_date.strftime("%d-%b-%Y")
    msg = EmailMessage()
    msg["Subject"] = f"Daily LeetCode Student Progress Report - {today_ist_date.strftime('%Y-%m-%d')}"
    msg["From"] = sender
    msg["To"] = RECEIVER_EMAIL
    msg.set_content(
        f"Dear Sir/Madam,\n\n"
        f"Please find attached the Daily LeetCode Student Progress Report for {date_str}.\n\n"
        f"The report contains the daily and overall LeetCode problem-solving progress of all students.\n\n"
        f"Regards,\nKalaimani"
    )

    with open(report_path, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=os.path.basename(report_path),
        )

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(sender, app_password)
        smtp.send_message(msg)
    print(f"Email sent to {RECEIVER_EMAIL}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--test", action="store_true", help="Test run: writes LeetCode_Daily_Report_TEST.xlsx, does not email")
    args = parser.parse_args()

    today_ist = datetime.now(IST).date()
    os.makedirs(REPORTS_DIR, exist_ok=True)

    print(f"Loading student master file: {STUDENTS_FILE}")
    df = load_students(STUDENTS_FILE)
    print(f"Loaded {len(df)} students.")

    results = process_students(df, today_ist)

    if args.test:
        out_path = os.path.join(REPORTS_DIR, "LeetCode_Daily_Report_TEST.xlsx")
    else:
        out_path = os.path.join(REPORTS_DIR, f"LeetCode_Daily_Report_{today_ist.isoformat()}.xlsx")

    build_report(results, today_ist, out_path)
    print(f"Report written to {out_path}")

    append_history(results, today_ist)

    checked_ok = sum(1 for r in results if r["today"] != "N/A")
    unavailable = sum(1 for r in results if r["today"] == "N/A")
    solved_today = sum(1 for r in results if isinstance(r["today"], int) and r["today"] > 0)
    total_today = sum(r["today"] for r in results if isinstance(r["today"], int))

    print("\n--- Summary ---")
    print(f"Total students: {len(results)}")
    print(f"Successfully checked: {checked_ok}")
    print(f"Unavailable / errored: {unavailable}")
    print(f"Students who solved >=1 problem today: {solved_today}")
    print(f"Total unique problems solved today (all students): {total_today}")

    if not args.test:
        send_email(out_path, today_ist)
    else:
        print("\nTest run -- email NOT sent. Run without --test for the real daily run.")


if __name__ == "__main__":
    main()
